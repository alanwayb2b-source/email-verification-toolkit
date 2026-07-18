import sys

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    print("Missing dependency: openpyxl")
    print("Install it with: pip install openpyxl")
    sys.exit(1)

from email_cleaner import clean_email, is_valid_email


def process_excel(input_file, output_file):
    """
    Clean emails, remove duplicates, and add validation status
    for Excel (.xlsx) files.
    """

    workbook = load_workbook(input_file)
    worksheet = workbook.active

    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in worksheet[1]
    ]

    email_column_index = None

    for index, header in enumerate(headers):
        if header.lower() == "email":
            email_column_index = index
            break

    if email_column_index is None:
        raise ValueError("Input Excel file must contain an 'email' column.")

    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "Cleaned Data"

    output_headers = headers + ["email_status"]
    output_sheet.append(output_headers)

    seen_emails = set()

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_data = list(row)

        email = clean_email(row_data[email_column_index])

        if not email or email in seen_emails:
            continue

        seen_emails.add(email)

        row_data[email_column_index] = email

        email_status = (
            "valid_format"
            if is_valid_email(email)
            else "invalid_format"
        )

        row_data.append(email_status)
        output_sheet.append(row_data)

    output_workbook.save(output_file)

    print(f"Done! Cleaned Excel file saved to: {output_file}")
    print(f"Unique email records processed: {len(seen_emails)}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(
            "Usage: python excel_cleaner.py "
            "input.xlsx output.xlsx"
        )
        sys.exit(1)

    process_excel(sys.argv[1], sys.argv[2])
